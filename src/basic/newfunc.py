# 全ての関数でタイトル付のCSVを読み込み、タイトル付のCSVを出力する
# csvin: 入力ファイル名
# csvout: 出力ファイル名
# csvinの列名は'col1', 'col2', 'col3'とする
# csvoutの列名は任意のものとする

import csv

# a_numberとanother_numberの積をcsvoutに出力
def kakezan(csvin, csvout):
    with open(csvin,  'r', newline='') as file_in,\
         open(csvout, 'w', newline='') as file_out:
        fieldnames = ['col1', 'col2', 'seki']
        reader = csv.DictReader(file_in, quoting=csv.QUOTE_NONNUMERIC)
        writer = csv.DictWriter(file_out, quoting=csv.QUOTE_NONNUMERIC,
                                fieldnames=fieldnames)
        writer.writeheader()
        for row in reader:
            ret1 = row['col1'] * row['col2']
            writer.writerow({
                'col1': row['col1'],
                'col2': row['col2'],
                'seki': ret1
            })

# a_numberより大きく一番近い素数と、a_numberより小さい素数の数をcsvoutに出力
def prime(csvin, csvout):
    # c:\ibi\apps\python\primelog.txtに現在日時を追記する
    from datetime import datetime
    with open('c:\\ibi\\apps\\python\\primelog.txt', 'a') as f:
        f.write(str(datetime.now()) + '\n')
        
    with open(csvin,  'r', newline='') as file_in,\
         open(csvout, 'w', newline='') as file_out:
        fieldnames = ['col1', 'next_prime', 'previous_prime']
        reader = csv.DictReader(file_in, quoting=csv.QUOTE_NONNUMERIC)
        writer = csv.DictWriter(file_out, quoting=csv.QUOTE_NONNUMERIC,
                                fieldnames=fieldnames)
        writer.writeheader()
        for row in reader:
            col1 = int(row['col1'])
            next_prime = col1 + 1
            while True:
                for i in range(2, next_prime):
                    if next_prime % i == 0:
                        break
                else:
                    break
                next_prime += 1
            previous_prime = col1 - 1
            while True:
                for i in range(2, previous_prime):
                    if previous_prime % i == 0:
                        break
                else:
                    break
                previous_prime -= 1
            # col1より小さい素数が見つからなかった場合、0を出力
            if previous_prime == 1:
                previous_prime = 0
            writer.writerow({'col1': col1,
                             'next_prime': next_prime,
                             'previous_prime': previous_prime})

# col1の中央値をcsvoutの全ての行に出力
def median(csvin, csvout):
    with open(csvin,  'r', newline='') as file_in,\
         open(csvout, 'w', newline='') as file_out:
        fieldnames = ['median']
        reader = csv.DictReader(file_in, quoting=csv.QUOTE_NONNUMERIC)
        writer = csv.DictWriter(file_out, quoting=csv.QUOTE_NONNUMERIC,
                                fieldnames=fieldnames)
        writer.writeheader()
        col1_list = []
        for row in reader:
            col1_list.append(row['col1'])
        col1_list.sort()
        median = col1_list[len(col1_list) // 2]
        for row in col1_list:   #入力と同じ数の行を出力する
            writer.writerow({'median': median})




# col1の偏差値をcsvoutの対応する行に出力
def hensachi(csvin, csvout):
    with open(csvin,  'r', newline='') as file_in,\
         open(csvout, 'w', newline='') as file_out:
        fieldnames = ['ret']
        reader = csv.DictReader(file_in, quoting=csv.QUOTE_NONNUMERIC)
        writer = csv.DictWriter(file_out, quoting=csv.QUOTE_NONNUMERIC,
                                fieldnames=fieldnames)
        writer.writeheader()
        col1_list = []
        for row in reader:
            col1_list.append(row['col1'])
        col1_list.sort()
        ave = sum(col1_list) / len(col1_list)
        std = (sum([(x - ave) ** 2 for x in col1_list]) / len(col1_list)) ** 0.5
        for row in col1_list:
            ret = 50 + 10 * (row - ave) / std
            writer.writerow({'ret': ret})

# csvinのcol1が年月,col2が売上金額、これを年月別売上グラフとしてExcelファイルに出力する関数
# 実績とともに売上の近似直線をグラフに描画する
# pip install openpyxl
def excelOut(csvin, csvout):
    with open(csvin,  'r', newline='') as file_in,\
         open(csvout, 'w', newline='') as file_out:
        fieldnames = ['ret']
        reader = csv.DictReader(file_in, quoting=csv.QUOTE_NONNUMERIC)
        writer = csv.DictWriter(file_out, quoting=csv.QUOTE_NONNUMERIC,
                                fieldnames=fieldnames)
        writer.writeheader()
        # Excelファイルに出力する処理を記述する
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '売上グラフ'
        ws['A1'] = '年月'
        ws['B1'] = '売上金額'
        row_num = 2
        for row in reader:
            ws[f'A{row_num}'] = row['col1']
            ws[f'B{row_num}'] = row['col2']
            row_num += 1
        chart = openpyxl.chart.LineChart()
        chart.title = '売上グラフ'
        chart.x_axis.title = '年月'
        chart.y_axis.title = '売上金額'
        chart.add_data(openpyxl.chart.Reference(ws, min_col=2, min_row=2, max_row=row_num))
        ws.add_chart(chart, 'D2')
        # c:\dev\python\売上グラフ.xlsxに保存する
        wb.save('c:\\dev\\python\\売上グラフ.xlsx')

        # csvinと同じ数の行の"ok"を出力する
        for row in reader:
            writer.writerow({'ret': 'ok'})


